import openpyxl

from Utilities.excelreader import readAllValues


def getDataExcel(sheetName):
    # workbook = openpyxl.load_workbook("../Excel/login.xlsx")
    # sheet = workbook[sheetName]
    # totalrows = sheet.max_row
    # totalcols = sheet.max_column
    # mainList = []
    #
    # for i in range(2, totalrows + 1):
    #     # Column data list
    #     dataList = []
    #     for j in range(1, totalcols + 1):
    #         data = sheet.cell(row=i, column=j).value
    #         dataList.insert(j, data)
    #     mainList.insert(i, dataList)
    # return mainList
    return readAllValues(sheetName)