import openpyxl

def getRowCount(sheetName):
    # Load the excel workbook
    workbook = openpyxl.load_workbook("../Excel/loginData.xlsx")
    # Load the excel workbook
    sheet = workbook[sheetName]
    # Total rows
    return sheet.max_row

def getColumnCount(sheetName):
    # Load the excel workbook
    workbook = openpyxl.load_workbook("../Excel/loginData.xlsx")
    # Load the excel workbook
    sheet = workbook[sheetName]
    # Total rows
    return sheet.max_column

def readAllValues(sheetName):
    # Load the excel workbook
    workbook = openpyxl.load_workbook("../Excel/loginData.xlsx")
    # Load the excel workbook
    sheet = workbook[sheetName]
    mainlist = []

    # To get all rows and all column data
    for i in range(2, sheet.max_row + 1):  # Rows
        # Column data list
        dataList = []
        for j in range(1, sheet.max_column + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainlist.insert(i,dataList)
    return mainlist


a = getRowCount("SendMail")
print(a)
b = getColumnCount("SendMail")
print(b)

listVal = readAllValues("SendMail")
print(listVal)
